# Import pandas
from openpyxl import load_workbook
from DataBase import DataBase
import pandas as pd
import xlsxwriter
import  datetime



class Writer:

    @staticmethod
    def stud():

        wb = load_workbook('./students.xlsx')

        sheet = wb.get_sheet_by_name('Sheet1')

        anotherSheet = wb.active
        num = 2
        B = 'B'
        C = 'C'
        D = 'D'
        E = 'E'
        DB = DataBase()
        while True:

            b = f"{B}{num}"
            c = f"{C}{num}"
            d = f"{D}{num}"
            e = f"{E}{num}"
            b2 = sheet[b].value
            c2 = sheet[c].value
            d2 = sheet[d].value
            e2 = sheet[e].value

            num += 1
            if b2 == None or c2 == None or d2 == None or e2 == None:
                break
            else:
                DB.student(b2, c2, d2, e2)

    @staticmethod
    def append(one,two,free,four,fife,six,seven,name,midlle_name,surname,user_num,balls):

        data = datetime.datetime.now()
        char_list = []
        data = str(data)

        for c in data:

            char_list.append(c)

        for i in range(7):

            del char_list[-1]

        data = char_list[8] + char_list[9] + '.' + char_list[5] + char_list[6] + '.' + char_list[0] + char_list[1] + \
               char_list[2] + char_list[3] + ' / ' + char_list[11] + char_list[12] + ':' + char_list[14] + char_list[15]

        if balls < 9:

            mark = 2
        else:

            if balls > 8 and balls < 13:

                mark = 3

            else:

                if balls > 12 and balls < 17:

                    mark = 4

                else:

                    if balls > 16:

                        mark = 5
        B = 'B'
        C = 'C'
        D = 'D'
        E = 'E'
        F = 'F'
        G = "G"
        H = 'H'
        Bcount = 2
        Ccount = 2
        Dcount = 2
        Ecount = 2
        Fcount = 2
        Gcount = 2
        Hcount = 2
        workbook = xlsxwriter.Workbook('mark.xlsx')

        worksheet = workbook.add_worksheet()
        worksheet.write("B1","name" )
        worksheet.write("C1", "midlle name")
        worksheet.write("D1", "surname")
        worksheet.write("E1", "number")
        worksheet.write("F1", "data")
        worksheet.write("G1", "balls")
        worksheet.write("H1", "mark")


        for i in range(len(one)):

            b2 = f"{B}{Bcount}"
            worksheet.write(b2,one[i])
            Bcount +=1

        b2 = f"{B}{Bcount}"
        worksheet.write(b2, name)



        for i in range(len(two)):

            c2 = f"{C}{Ccount}"
            worksheet.write(c2,two[i])
            Ccount +=1

        c2 = f"{C}{Ccount}"
        worksheet.write(c2, midlle_name)

        for i in range(len(free)):

            d2 = f"{D}{Dcount}"
            worksheet.write(d2,free[i])
            Dcount +=1

        d2 = f"{D}{Dcount}"
        worksheet.write(d2, surname)

        for i in range(len(four)):

            e2 = f"{E}{Ecount}"
            worksheet.write(e2,four[i])
            Ecount +=1

        e2 = f"{E}{Ecount}"
        worksheet.write(e2, user_num)

        for i in range(len(fife)):

            f2 = f"{F}{Fcount}"
            worksheet.write(f2,fife[i])
            Fcount +=1

        f2 = f"{F}{Fcount}"
        worksheet.write(f2, data)

        for i in range(len(six)):

            g2 = f"{G}{Gcount}"
            worksheet.write(g2,six[i])
            Gcount +=1

        g2 = f"{G}{Gcount}"
        worksheet.write(g2, balls)

        for i in range(len(seven)):

            h2 = f"{H}{Hcount}"
            worksheet.write(h2,seven[i])
            Hcount +=1

        h2 = f"{H}{Hcount}"
        worksheet.write(h2, mark)





        workbook.close()

    @staticmethod

    def wirite(name,midlle_name,surname,user_num,balls):
        wb = load_workbook('mark.xlsx')

        sheet = wb.get_sheet_by_name('Sheet1')

        anotherSheet = wb.active
        num = 2
        B = 'B'
        C = 'C'
        D = 'D'
        E = 'E'
        F = 'F'
        G = "G"
        H = 'H'
        num = 2

        bl = []
        cl = []
        dl = []
        el = []
        fl = []
        gl = []
        hl = []

        while True:

            b = f"{B}{num}"
            c = f"{C}{num}"
            d = f"{D}{num}"
            e = f"{E}{num}"
            f = f"{F}{num}"
            g = f"{G}{num}"
            h = f"{H}{num}"
            b2 = sheet[b].value
            c2 = sheet[c].value
            d2 = sheet[d].value
            e2 = sheet[e].value
            f2 = sheet[f].value
            g2 = sheet[g].value
            h2 = sheet[h].value
            num = num + 1

            if b2 == None: #or c2 == None or d2 == None or e2 == None or f2 == None or g2 == None or h2 == None:
                break


            else:

                bl.append(b2)
                cl.append(c2)
                dl.append(d2)
                el.append(e2)
                fl.append(f2)
                gl.append(g2)
                hl.append(h2)

        Writer.append(bl,cl,dl,el,fl,gl,hl,name,midlle_name,surname,user_num,balls)





